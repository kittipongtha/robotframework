import  openpyxl

wb = openpyxl.load_workbook('C:/Users/kittipong.t/Desktop/practice/practice-robotframework/test_data/test_scenario_krungsri.xlsx')
sheet = wb['testdata-t001']

test_case_name = ['test_case_name']
name = ['name']
surname = ['surname']
identification_number = ['identification_number']
mobile = ['mobile']
email = ['email']
expected = ['expected']
occupation = ['occupation']
salary = ['salary']

def test(choose):
    if test_case_name[0] == choose:
        test_case_name.remove('test_case_name')
        choose = test_case_name
        get_data.get_test_case_name()
    elif name[0] == choose:
        name.remove('name')
        choose = name
        get_data.get_name()
    elif surname[0] == choose:
        surname.remove('surname')
        choose = surname
        get_data.get_surname()
    elif identification_number[0] == choose:
        identification_number.remove('identification_number')
        choose = identification_number
        get_data.get_identification_number
    elif mobile[0] == choose:
        mobile.remove('mobile')
        choose = mobile
        get_data.get_mobile()
    elif email[0] == choose:
        email.remove('email')
        choose = email
        get_data.get_email()
    elif expected[0] == choose:
        expected.remove('expected')
        choose = expected
        get_data.get_expected()
    elif occupation[0] == choose:
        occupation.remove('occupation')
        choose = occupation
        get_data.get_occupation()
    elif salary[0] == choose:
        salary.remove('salary')
        choose = salary
        get_data.get_salary()

    return  choose

class   get_data:
    def get_test_case_name():
        for row in range(2, sheet.max_row + 1):
            test_case_name.append(sheet['A' + str(row)].value)
        return  test_case_name

    def get_name():
        for row in range(2, sheet.max_row + 1):
            name.append(sheet['B' + str(row)].value)
        return  name
    
    def get_surname():
        for row in range(2, sheet.max_row + 1):
            surname.append(sheet['C' + str(row)].value)
        return  surname

    def get_identification_number():
        for row in range(2, sheet.max_row + 1):
            identification_number.append(sheet['C' + str(row)].value)
        return  identification_number

    def get_mobile():
        for row in range(2, sheet.max_row + 1):
            mobile.append(sheet['E' + str(row)].value)
        return  mobile

    def get_email():
        for row in range(2, sheet.max_row + 1):
            email.append(sheet['F' + str(row)].value)
        return  email

    def get_expected():
        for row in range(2, sheet.max_row + 1):
            expected.append(sheet['G' + str(row)].value)
        return  expected

    def get_occupation():
        for row in range(2, sheet.max_row + 1):
            occupation.append(sheet['H' + str(row)].value)
        return  occupation

    def get_salary():
        for row in range(2, sheet.max_row + 1):
            salary.append(sheet['I' + str(row)].value)
        return  salary

# name = test('name')

# print(name)